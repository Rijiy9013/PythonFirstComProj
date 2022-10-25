import time
import tkinter as tk
from tkinter import *
import random
from time import *
from openpyxl import Workbook
from openpyxl import load_workbook


class PsyProgram:
    window = tk.Tk()
    size_image = 2
    start_time = 0
    end_time = 0
    numb_instr = 0
    numb_image = 0
    randPosOrNeg = 0
    blockNumber = 0
    wait_till_test = tk.IntVar()
    wait = tk.IntVar()
    wait_error = tk.IntVar()
    participant_number = 1
    participant_age = ''
    job = ""
    participant_work = 0
    participant_name = ''
    firstSectOver = False
    trainOver = False
    check_pref = False
    loadExcel = False
    wasError = False
    instr_name = "instruction\\"
    photos_name = "photos\\"
    test_name = "test_photos\\"
    numb_test_pic = 2
    counter = 0
    checkAorB = 0  # 0 - полож, 1 - нег
    counterA = 0
    counterB = 0
    list_mainFP_order = []
    list_mainFN_order = []
    list_mainSP_order = []
    list_mainSN_order = []
    list_testFP_order = []
    list_testFN_order = []
    list_testSP_order = []
    list_testSN_order = []
    works = ['Руководитель предприятия, учреждения', 'Инженерно-технический работник',
             'Служащий аппарата управления предприятия, учреждения',
             'Военнослужащий, работник правоохранительных органов, юстиции', 'Рабочий', 'Домохозяйка',
             'Предприниматель', 'Работник сферы обслуживания', 'Студент, учащийся',
             'Безработный, временно неработающий',
             'Представитель интеллигенции (образование, здравоохранение, культура)', 'Пенсионер',
             'Другой вид деятельности (укажите)']

    def __init__(self):
        self.window.title('Тестирование')
        self.window.geometry("1500x1300")
        self.mainTimeList = []
        self.trainTimeList = []
        self.valenceTrainList = []
        self.valenceMainList = []
        self.movementTrainList = []
        self.movementMainList = []
        self.accTrainList = []
        self.accMainList = []
        self.errorTrainList = []
        self.errorMainList = []
        self.mainPhotosNumb = []
        self.testPhotosNumb = []
        self.file1 = Workbook()
        self.menubar = tk.Menu(self.window)
        self.window.config(menu=self.menubar)
        settings_menu = tk.Menu(self.menubar, tearoff=0)
        settings_menu.add_command(label='Старт', command=self.start_test)
        settings_menu.add_command(label='Изображения', command=self.create_settings_win)
        self.menubar.add_cascade(label='Настройки', menu=settings_menu)
        self.canvas = Canvas(self.window, width=1200, height=1000)
        self.canvas.pack()
        self.background_obj = PhotoImage(file=r'background\startBackgr.png')
        self.end_image = PhotoImage(file=r'background\end.png')
        self.BGpreFix = PhotoImage(file=r'background\prefixation.png')  # до нажатия на h
        self.BGfix = PhotoImage(file=r'background\fixation.png')  # в момент нажатия
        self.BGfrw = PhotoImage(file=r'background\BackgrFrw.png')
        self.BGback = PhotoImage(file=r'background\BackgrBack.png')
        self.canvas.create_image(20, 20, anchor="nw", image=self.background_obj)
        self.error = PhotoImage(file=r'photos\error.png')
        self.instr = PhotoImage()
        self.instr_list = []
        self.photosA_list = []
        self.photosB_list = []
        self.photosA_test = []
        self.photosB_test = []
        self.photosAFrw_list = []
        self.photosABck_list = []
        self.photosBFrw_list = []
        self.photosBBck_list = []
        self.photosAFrw_test = []
        self.photosABck_test = []
        self.photosBFrw_test = []
        self.photosBBck_test = []
        self.photosA = PhotoImage()
        self.photosB = PhotoImage()
        self.photosAFrw = PhotoImage()
        self.photosABck = PhotoImage()
        self.photosBFrw = PhotoImage()
        self.photosBBck = PhotoImage()

    def start(self):
        PsyProgram.window.mainloop()

    def error_no_excel(self):
        error_win = tk.Toplevel(self.window)
        error_win.wm_title('Ошибка')
        error_win.geometry("375x100")
        tk.Label(error_win, text='Проверьте имя Excel файла (должен быть Results.xlsx)').grid(row=0, column=1)

    def load_pictures(self):
        if self.loadExcel:
            try:
                self.file1 = load_workbook('Results.xlsx')
            except FileNotFoundError:
                self.error_no_excel()
        for i in range(10):
            self.instr = PhotoImage(file=self.instr_name + str(i+1)+".png")
            self.instr_list.append(self.instr)
        for i in range(self.size_image):
            self.photosA = PhotoImage(file=self.photos_name + str(i+1)+"A.png")
            self.photosABck = PhotoImage(file=self.photos_name + str(i+1) + "Abck.png")
            self.photosAFrw = PhotoImage(file=self.photos_name + str(i+1) + "Afrw.png")
            self.photosA_list.append(self.photosA)
            self.photosABck_list.append(self.photosABck)
            self.photosAFrw_list.append(self.photosAFrw)
            self.photosB = PhotoImage(file=self.photos_name + str(i+1) + "B.png")
            self.photosBBck = PhotoImage(file=self.photos_name + str(i+1) + "Bbck.png")
            self.photosBFrw = PhotoImage(file=self.photos_name + str(i+1) + "Bfrw.png")
            self.photosB_list.append(self.photosB)
            self.photosBBck_list.append(self.photosBBck)
            self.photosBFrw_list.append(self.photosBFrw)
        for i in range(self.numb_test_pic):
            self.photosA = PhotoImage(file=self.test_name + str(i + 1) + "A.png")
            self.photosABck = PhotoImage(file=self.test_name + str(i + 1) + "Abck.png")
            self.photosAFrw = PhotoImage(file=self.test_name + str(i + 1) + "Afrw.png")
            self.photosA_test.append(self.photosA)
            self.photosABck_test.append(self.photosABck)
            self.photosAFrw_test.append(self.photosAFrw)
            self.photosB = PhotoImage(file=self.test_name + str(i + 1) + "B.png")
            self.photosBBck = PhotoImage(file=self.test_name + str(i + 1) + "Bbck.png")
            self.photosBFrw = PhotoImage(file=self.test_name + str(i + 1) + "Bfrw.png")
            self.photosB_test.append(self.photosB)
            self.photosBBck_test.append(self.photosBBck)
            self.photosBFrw_test.append(self.photosBFrw)

    def start_test(self):
        if self.size_image != 0 and self.numb_test_pic != 0:
            self.counterA = 0
            self.counterB = 0
            self.counter = 0
            self.numb_instr = 0
            self.blockNumber = 0
            self.firstSectOver = False
            self.trainOver = False
            self.check_pref = False
            self.list_clear()
            self.randPosOrNeg = random.randint(0, 1)  # 0 - сначала пол на Y, 1 - наоборот, сначала отр на Y
            self.get_data()
            self.print_instruction()
        else:
            self.print_error()

    def list_clear(self):
        self.errorMainList.clear()
        self.errorTrainList.clear()
        self.movementMainList.clear()
        self.movementTrainList.clear()
        self.valenceMainList.clear()
        self.valenceTrainList.clear()
        self.mainTimeList.clear()
        self.trainTimeList.clear()
        self.accMainList.clear()
        self.accTrainList.clear()
        self.mainPhotosNumb.clear()
        self.testPhotosNumb.clear()

    def create_settings_win(self):
        win_settings = tk.Toplevel(self.window)
        check = tk.IntVar()
        win_settings.wm_title('Количество изображений')
        win_settings.geometry("500x350")
        tk.Label(win_settings, text='Фото для основного этапа: ').grid(row=0, column=0)
        tk.Label(win_settings, text='Фото для этапа обучения: ').grid(row=1, column=0)
        tk.Label(win_settings, text='Номер испытуемого: ').grid(row=2, column=0)
        var = tk.BooleanVar()
        var.set(False)
        image_entry4 = tk.Checkbutton(win_settings, text='Открыть существующий Excel?', variable=var, onvalue=1,
                                      offvalue=0)
        image_entry4.grid(row=3, column=1, padx=20, pady=20)
        image_entry1 = tk.Entry(win_settings)
        image_entry1.grid(row=0, column=1, padx=20, pady=20)
        image_entry2 = tk.Entry(win_settings)
        image_entry2.grid(row=1, column=1, padx=20, pady=20)
        image_entry3 = tk.Entry(win_settings)
        image_entry3.grid(row=2, column=1, padx=20, pady=20)
        image_entry1.insert(0, str(PsyProgram.size_image))
        image_entry2.insert(0, str(PsyProgram.numb_test_pic))
        image_entry3.insert(0, str(PsyProgram.participant_number))
        tk.Button(win_settings, text='Применить', command=lambda: check.set(1)).grid(
            row=4, column=0, columnspan=2)
        win_settings.wait_variable(check)
        self.size_image = int(image_entry1.get())
        self.numb_test_pic = int(image_entry2.get())
        self.participant_number = int(image_entry3.get())
        self.loadExcel = var.get()
        if self.size_image < self.numb_test_pic:
            self.error_wrong_input2()
            win_settings.destroy()
        try:
            self.load_pictures()
        except TclError:
            self.error_wrong_input()
        win_settings.destroy()

    def change_settings(self, image: tk.Entry):  # 72 - h, 89 - y, 78 - n
        PsyProgram.size_image = int(image.get())

    def print_instruction(self):
        self.window.bind("<KeyPress-space>", self.next_instr)
        self.canvas.delete("all")
        if not self.firstSectOver:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[4], tags="fifth")
            if self.randPosOrNeg == 0:
                self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[3], tags="fourth")
            else:
                self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[7], tags="fourth")
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[2], tags="third")
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[1], tags="second")
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[0], tags="first")

    def next_instr(self, event):
        list_instr = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh', 'eight', 'ninth', 'tenth']
        if event.keysym == 'space':
            self.canvas.delete(list_instr[self.numb_instr])
            self.numb_instr += 1
            if self.numb_instr == 5:
                self.numb_instr = 0
                self.test_play()
        else:
            return False

    def another_instr(self):
        self.canvas.delete("all")
        self.randPosOrNeg = int(not self.randPosOrNeg)
        self.window.bind("<KeyPress-space>", self.skip_an_instr)
        self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[4], tags="fifth")
        if self.randPosOrNeg == 0:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[3], tags="fourth")
        else:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[7], tags="fourth")
        # self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[2], tags="third")
        # self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[1], tags="second")

    def skip_an_instr(self, event):
        list_instr = ['fourth', 'fifth', 'sixth', 'seventh', 'eight', 'ninth', 'tenth']
        if event.keysym == 'space':
            self.canvas.delete(list_instr[self.numb_instr])
            self.numb_instr += 1
            if self.numb_instr == 2:
                self.test_play()

    def show_error(self):
        self.canvas.delete("all")
        self.canvas.create_image(20, 20, anchor="nw", image=self.error, tag='error')
        if self.checkAorB == 0:
            self.counterA += 1
        elif self.checkAorB == 1:
            self.counterB += 1
        self.pause()

    def test_play(self):
        self.trainOver = False
        if self.firstSectOver:
            self.list_testSP_order = random.sample(range(0, self.numb_test_pic), self.numb_test_pic)
            self.list_testSN_order = random.sample(range(0, self.numb_test_pic), self.numb_test_pic)
        else:
            self.list_testFP_order = random.sample(range(0, self.numb_test_pic), self.numb_test_pic)
            self.list_testFN_order = random.sample(range(0, self.numb_test_pic), self.numb_test_pic)
        self.window.unbind("<KeyPress-space>")
        self.canvas.delete("all")
        self.show_test_pictures()

    def show_test_pictures(self):
        if self.counterA + self.counterB < 2 * self.numb_test_pic:
            self.wasError = False
            if not self.check_pref:
                self.window.bind("<KeyPress-h>", self.show_bgs)
                self.show_prefix()
            else:
                if self.counterA == self.numb_test_pic:
                    randNum = 2
                elif self.counterB == self.numb_test_pic:
                    randNum = 1
                else:
                    randNum = random.randint(1, 2)  # 1 - a, 2 - b
                self.start_time = time()
                if randNum == 1:
                    if self.firstSectOver:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosA_test[self.list_testSP_order[self.counterA]])
                    else:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosA_test[self.list_testFP_order[self.counterA]])
                    self.checkAorB = 0
                    self.testPhotosNumb.append(self.counterA)
                else:
                    if self.firstSectOver:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosB_test[self.list_testSN_order[self.counterB]])
                    else:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosB_test[self.list_testFN_order[self.counterB]])
                    self.checkAorB = 1
                    self.testPhotosNumb.append(self.counterB)
        else:
            self.counterA = self.counterB = 0
            self.check_pref = False
            self.trainOver = True
            self.main_play()
        self.check_pref = False

    def test_resY(self, arg):
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.end_time = time()
        self.canvas.delete("all")
        if self.randPosOrNeg == 0 and self.checkAorB == 0:  # приближаем позитивные
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosAFrw_test[self.list_testSP_order[self.counterA]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosAFrw_test[self.list_testFP_order[self.counterA]])
            self.counterA += 1
            self.valenceTrainList.append('1')
            self.errorTrainList.append('1')
        elif self.randPosOrNeg == 1 and self.checkAorB == 1:  # приближаем негативные
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBFrw_test[self.list_testSN_order[self.counterB]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBFrw_test[self.list_testFN_order[self.counterB]])
            self.counterB += 1
            self.valenceTrainList.append('2')
            self.errorTrainList.append('1')
        else:
            self.wasError = True
            self.show_error()
            self.valenceTrainList.append(str(self.checkAorB + 1))
            self.errorTrainList.append('2')
        self.trainTimeList.append(self.end_time - self.start_time)
        self.movementTrainList.append('1')
        self.counter += 1
        self.pause()

    def test_resN(self, arg):
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.end_time = time()
        self.canvas.delete("all")
        if self.randPosOrNeg == 1 and self.checkAorB == 0:  # отдаляем позитивные
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosABck_test[self.list_testSP_order[self.counterA]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosABck_test[self.list_testFP_order[self.counterA]])
            self.counterA += 1
            self.valenceTrainList.append('1')
            self.errorTrainList.append('1')
        elif self.randPosOrNeg == 0 and self.checkAorB == 1:  # отдаляем негативные
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBBck_test[self.list_testSN_order[self.counterB]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBBck_test[self.list_testFN_order[self.counterB]])
            self.counterB += 1
            self.valenceTrainList.append('2')
            self.errorTrainList.append('1')
        else:
            self.wasError = True
            self.show_error()
            self.valenceTrainList.append(str(self.checkAorB + 1))
            self.errorTrainList.append('2')
        self.trainTimeList.append(self.end_time - self.start_time)
        self.movementTrainList.append('2')
        self.counter += 1
        self.pause()

    def del_now(self):
        self.canvas.delete("all")
        if not self.trainOver:
            self.window.bind("<KeyPress-y>", self.test_resY)
            self.window.bind("<KeyPress-n>", self.test_resN)
            self.show_test_pictures()
        else:
            self.window.bind("<KeyPress-y>", self.resY)
            self.window.bind("<KeyPress-n>", self.resN)
            self.show_main_pictures()

    def pause(self):
        self.window.after(2000, self.del_now)

    def pause_bg(self):
        self.window.after(1500, self.del_now)

    def show_bgs(self, arg):
        self.window.unbind("<KeyPress-h>")
        self.check_pref = True
        self.canvas.delete("all")
        self.canvas.create_image(20, 20, anchor="nw", image=self.BGfix)
        self.pause_bg()

    def save_results(self):
        if not self.loadExcel:
            wsMain = self.file1.active
            wsMain.title = "Технические условия"
            wsMain["A1"] = "Номер фото (берется из имени png)"
            wsMain["B1"] = "Валентность (1 - полож, 2 - отр)"
            wsMain["C1"] = "Движение (1 - приближение, 2 - отдаление)"
            wsMain["D1"] = "Ошибка (1 - правильно, 2 - неправильно)"
            wsMain["E1"] = "Время"
            wsMain["F1"] = "Порядок блока( 1 - приближаем положительные, 2 - приближаем отрицательные)"
            wsMain["G1"] = "Номер испытуемого"
            wsMain["H1"] = "Имя испытуемого"
            wsMain["I1"] = "Возраст испытуемого"
            wsMain["J1"] = "Работа испытуемого"
            wsMain["K1"] = "Т - обучение, затем идет номер испытуемого. Пример: Т1 - обучение 1-ого испытуемого"
            wsMain["L1"] = "O - основной блок, затем идет номер блока и номер испытуемого. Пример: О12" \
                           " - первой основной блок для второго испытуемого, соотвественно, О22 - 2-ой основной блок" \
                           "для 2-ого испытуемого"
        row = 2
        ws1 = self.file1.create_sheet("T" + str(self.participant_number))
        ws2 = self.file1.create_sheet("O1" + str(self.participant_number))
        ws3 = self.file1.create_sheet("O2" + str(self.participant_number))
        ws1["A1"] = "Фото"
        ws1["B1"] = "Вал"
        ws1["C1"] = "Движ"
        ws1["D1"] = "Ошибка"
        ws1["E1"] = "Время"
        ws1["F1"] = "Порядок"
        ws1["G1"] = "Номер"
        ws1["H1"] = "Имя"
        ws1["I1"] = "Возраст"
        ws1["J1"] = "Работа"
        ws2["A1"] = "Фото"
        ws2["B1"] = "Вал"
        ws2["C1"] = "Движ"
        ws2["D1"] = "Ошибка"
        ws2["E1"] = "Время"
        ws2["F1"] = "Порядок"
        ws2["G1"] = "Номер"
        ws2["H1"] = "Имя"
        ws2["I1"] = "Возраст"
        ws2["J1"] = "Работа"
        ws3["A1"] = "Фото"
        ws3["B1"] = "Вал"
        ws3["C1"] = "Движ"
        ws3["D1"] = "Ошибка"
        ws3["E1"] = "Время"
        ws3["F1"] = "Порядок"
        ws3["G1"] = "Номер"
        ws3["H1"] = "Имя"
        ws3["I1"] = "Возраст"
        ws3["J1"] = "Работа"
        counterp = countern = 0
        for i in range(len(self.trainTimeList) // 2):
            ws1["B" + str(row)] = str(self.valenceTrainList[i])
            if (self.valenceTrainList[i] == '1'):
                ws1["A" + str(row)] = str(self.list_testFP_order[counterp] + 1)
                counterp += 1
            else:
                ws1["A" + str(row)] = str(self.list_testFN_order[countern] + 1)
                countern += 1
            ws1["C" + str(row)] = str(self.movementTrainList[i])
            ws1["D" + str(row)] = str(self.errorTrainList[i])
            ws1["E" + str(row)] = str(self.trainTimeList[i])
            ws1["F" + str(row)] = str(int(not self.randPosOrNeg) + 1)
            ws1["G" + str(row)] = str(self.participant_number)
            ws1["H" + str(row)] = str(self.participant_name)
            ws1["I" + str(row)] = str(self.participant_age)
            if self.participant_work == len(self.works) - 1:
                ws1["J" + str(row)] = str(self.job)
            else:
                ws1["J" + str(row)] = str(self.works[self.participant_work])
            row += 1
        counterp = countern = 0
        for i in range(len(self.trainTimeList) // 2, len(self.trainTimeList)):
            ws1["B" + str(row)] = str(self.valenceTrainList[i])
            if (self.valenceTrainList[i] == '1'):
                ws1["A" + str(row)] = str(self.list_testSP_order[counterp] + 1)
                counterp += 1
            else:
                ws1["A" + str(row)] = str(self.list_testSN_order[countern] + 1)
                countern += 1
            ws1["C" + str(row)] = str(self.movementTrainList[i])
            ws1["D" + str(row)] = str(self.errorTrainList[i])
            ws1["E" + str(row)] = str(self.trainTimeList[i])
            ws1["F" + str(row)] = str(self.randPosOrNeg + 1)
            ws1["G" + str(row)] = str(self.participant_number)
            ws1["H" + str(row)] = str(self.participant_name)
            ws1["I" + str(row)] = str(self.participant_age)
            if self.participant_work == len(self.works) - 1:
                ws1["J" + str(row)] = str(self.job)
            else:
                ws1["J" + str(row)] = str(self.works[self.participant_work])
            row += 1
        row = 2
        countern = counterp = 0
        for i in range(len(self.mainTimeList) // 2):
            ws2["B" + str(row)] = str(self.valenceMainList[i])
            if (self.valenceMainList[i] == '1'):
                ws2["A" + str(row)] = str(self.list_mainFP_order[counterp] + 1)
                counterp += 1
            else:
                ws2["A" + str(row)] = str(self.list_mainFN_order[countern] + 1)
                countern += 1
            ws2["C" + str(row)] = str(self.movementMainList[i])
            ws2["D" + str(row)] = str(self.errorMainList[i])
            ws2["E" + str(row)] = str(self.mainTimeList[i])
            ws2["F" + str(row)] = str(int(not self.randPosOrNeg) + 1)
            ws2["G" + str(row)] = str(self.participant_number)
            ws2["H" + str(row)] = str(self.participant_name)
            ws2["I" + str(row)] = str(self.participant_age)
            if self.participant_work == len(self.works) - 1:
                ws2["J" + str(row)] = str(self.job)
            else:
                ws2["J" + str(row)] = str(self.works[self.participant_work])
            row += 1
        row = 2
        countern = counterp = 0
        for i in range(len(self.mainTimeList) // 2, len(self.mainTimeList)):
            ws3["B" + str(row)] = str(self.valenceMainList[i])
            if (self.valenceMainList[i] == '1'):
                ws3["A" + str(row)] = str(self.list_mainSP_order[counterp] + 1)
                counterp += 1
            else:
                ws3["A" + str(row)] = str(self.list_mainSN_order[countern] + 1)
                countern += 1
            ws3["C" + str(row)] = str(self.movementMainList[i])
            ws3["D" + str(row)] = str(self.errorMainList[i])
            ws3["E" + str(row)] = str(self.mainTimeList[i])
            ws3["F" + str(row)] = str(self.randPosOrNeg + 1)
            ws3["G" + str(row)] = str(self.participant_number)
            ws3["H" + str(row)] = str(self.participant_name)
            ws3["I" + str(row)] = str(self.participant_age)
            if self.participant_work == len(self.works) - 1:
                ws3["J" + str(row)] = str(self.job)
            else:
                ws3["J" + str(row)] = str(self.works[self.participant_work])
            row += 1
        try:
            self.file1.save("Results.xlsx")
        except PermissionError:
            error_win = tk.Toplevel(self.window)
            error_win.wm_title('Ошибка')
            error_win.geometry("650x100")
            tk.Label(error_win, text='Файл с результатами открыт, закройте его. Данные сохранены'
                                     'в ' + "ErrorResults" + str(self.participant_number) + ".xlsx").grid(row=0,
                                                                                                          column=1)
            self.file1.save("ErrorResults" + str(self.participant_number) + ".xlsx")

    def show_prefix(self):
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.canvas.delete("all")
        self.canvas.create_image(20, 20, anchor="nw", image=self.BGpreFix, tags='prefix')

    def main_play(self):
        self.trainOver = True
        self.counterA = self.counterB = 0
        if self.firstSectOver:
            self.list_mainSP_order = random.sample(range(0, self.size_image), self.size_image)
            self.list_mainSN_order = random.sample(range(0, self.size_image), self.size_image)
        else:
            self.list_mainFP_order = random.sample(range(0, self.size_image), self.size_image)
            self.list_mainFN_order = random.sample(range(0, self.size_image), self.size_image)
        self.canvas.delete("all")
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.window.unbind("<KeyPress-h>")
        self.show_reminder()

    def show_reminder(self):
        if self.wasError:
            return
        self.window.bind("<KeyPress-space>", self.skip_reminder)
        self.canvas.delete("all")
        if self.randPosOrNeg == 0:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[5])
        else:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[8])
        self.window.wait_variable(self.wait_error)

    def skip_reminder(self, arg):
        self.window.unbind("<KeyPress-space>")
        self.canvas.delete("all")
        self.show_main_pictures()

    def show_main_pictures(self):
        if self.wasError:
            self.wasError = False
            self.show_reminder()
        self.canvas.delete("all")
        if self.counterA + self.counterB != 2*self.size_image:
            if not self.check_pref:
                self.window.bind("<KeyPress-h>", self.show_bgs)
                self.show_prefix()
            else:
                if self.counterA == self.size_image:
                    randNum = 2
                elif self.counterB == self.size_image:
                    randNum = 1
                else:
                    randNum = random.randint(1, 2)  # 1 - pos, 2 - neg
                self.start_time = time()
                if randNum == 1:
                    if self.firstSectOver:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosA_list[self.list_mainSP_order[self.counterA]])
                    else:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosA_list[self.list_mainFP_order[self.counterA]])
                    self.checkAorB = 0
                    self.mainPhotosNumb.append(self.counterA)
                else:
                    if self.firstSectOver:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosB_list[self.list_mainSN_order[self.counterB]])
                    else:
                        self.canvas.create_image(20, 20, anchor="nw",
                                                 image=self.photosB_list[self.list_mainFN_order[self.counterB]])
                    self.checkAorB = 1
                    self.mainPhotosNumb.append(self.counterB)
            self.check_pref = False
        else:
            self.check_end()

    def resY(self, arg):
        self.end_time = time()
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.canvas.delete("all")
        if self.checkAorB == 0:
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosAFrw_list[self.list_mainSP_order[self.counterA]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosAFrw_list[self.list_mainFP_order[self.counterA]])
            self.counterA += 1
            self.valenceMainList.append('1')
        else:
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBFrw_list[self.list_mainSN_order[self.counterB]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBFrw_list[self.list_mainFN_order[self.counterB]])
            self.counterB += 1
            self.valenceMainList.append('2')
        if (self.checkAorB == 0 and self.randPosOrNeg == 0) or (self.checkAorB == 1 and self.randPosOrNeg == 1):
            self.errorMainList.append('1')
        else:
            self.errorMainList.append('2')
        self.mainTimeList.append(self.end_time - self.start_time)
        self.movementMainList.append('1')
        self.counter += 1
        self.pause()

    def resN(self, arg):
        self.end_time = time()
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.canvas.delete("all")
        if self.checkAorB == 0:
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosABck_list[self.list_mainSP_order[self.counterA]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosABck_list[self.list_mainFP_order[self.counterA]])
            self.counterA += 1
            self.valenceMainList.append('1')
        else:
            if self.firstSectOver:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBBck_list[self.list_mainSN_order[self.counterB]])
            else:
                self.canvas.create_image(20, 20, anchor="nw",
                                         image=self.photosBBck_list[self.list_mainFN_order[self.counterB]])
            self.counterB += 1
            self.valenceMainList.append('2')
        if (self.checkAorB == 1 and self.randPosOrNeg == 0) or (self.checkAorB == 0 and self.randPosOrNeg == 1):
            self.errorMainList.append('1')
        else:
            self.errorMainList.append('2')
        self.mainTimeList.append(self.end_time - self.start_time)
        self.movementMainList.append('2')
        self.counter += 1
        self.pause()

    def check_end(self):
        if not self.firstSectOver:
            self.window.unbind("<KeyPress-y>")
            self.window.unbind("<KeyPress-n>")
            self.window.unbind("<KeyPress-h>")
            self.show_first_sec_over()
        else:
            self.save_results()
            self.participant_number += 1
            self.end()

    def end(self):
        self.canvas.delete("all")
        self.canvas.create_image(20, 20, anchor="nw", image=self.end_image)

    def get_data(self):
        wait_var = tk.IntVar()
        win_data = tk.Toplevel(self.window)
        win_data.wm_title('Параметры')
        win_data.geometry("450x200")
        tk.Label(win_data, text='Возраст:').grid(row=0, column=0)
        age_entry = tk.Entry(win_data)
        age_entry.grid(row=0, column=1, padx=20, pady=20)
        tk.Label(win_data, text='Имя:').grid(row=1, column=0)
        name_entry = tk.Entry(win_data)
        name_entry.grid(row=1, column=1, padx=20, pady=20)
        button = tk.Button(win_data, text='Применить', command=lambda: wait_var.set(1))
        button.grid(row=3, column=0, columnspan=2)
        button.wait_variable(wait_var)
        self.participant_age = age_entry.get()
        self.participant_name = name_entry.get()
        win_data.destroy()
        self.get_job()

    def get_job(self):
        win_data = tk.Toplevel(self.window)
        win_data.wm_title('Параметры')
        win_data.geometry("700x330")
        listb = tk.Listbox(win_data, width=70, height=15)
        listb.pack()
        for i, elem in enumerate(self.works):
            listb.insert(i, elem)
        tk.Button(win_data, text='Применить',
                  command=lambda: self.wait_till_test.set(1)).pack(pady=20)
        win_data.wait_variable(self.wait_till_test)
        self.participant_work = listb.curselection()[0]
        win_data.destroy()
        if self.participant_work == len(self.works) - 1:
            self.get_name_job()

    def get_name_job(self):
        win = tk.Toplevel(self.window)
        win.wm_title('Параметры')
        win.geometry("330x130")
        tk.Label(win, text='Профессия: ').grid(row=0, column=0)
        job_entry = tk.Entry(win)
        job_entry.grid(row=0, column=1, padx=20, pady=20)
        wait_var = tk.IntVar()
        button = tk.Button(win, text='Применить', command=lambda: wait_var.set(1)).grid(row=1, column=0, columnspan=2)
        win.wait_variable(wait_var)
        self.job = str(job_entry.get())
        win.destroy()

    def show_first_sec_over(self):
        self.canvas.delete("all")
        self.window.bind("<KeyPress-space>", self.skip_sec_over)
        self.window.unbind("<KeyPress-y>")
        self.window.unbind("<KeyPress-n>")
        self.window.unbind("<KeyPress-h>")
        self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[6])

    def skip_sec_over(self, arg):
        self.canvas.delete("all")
        self.window.unbind("<KeyPress-space>")
        self.counterA = self.counterB = 0
        self.firstSectOver = True
        self.check_pref = self.trainOver = False
        self.another_instr()

    def print_error(self):
        error_win = tk.Toplevel(self.window)
        error_win.wm_title('Ошибка')
        error_win.geometry("350x100")
        tk.Label(error_win, text='Укажите количество картинок!').grid(row=0, column=1)

    def error_wrong_input(self):
        error_win = tk.Toplevel(self.window)
        error_win.wm_title('Ошибка')
        error_win.geometry("450x200")
        tk.Label(error_win, text='Такого количества картинок для основного этапа нет в папке photos').grid(row=0, column=1)
        tk.Label(error_win, text='Или же нет такого количества картинок для обучения в папке test_photos').grid(row=1,
                                                                                                           column=1)

    def error_wrong_input2(self):
        error_win = tk.Toplevel(self.window)
        error_win.wm_title('Ошибка')
        error_win.geometry("575x100")
        tk.Label(error_win, text='Количество картинок для обучения больше количества для основного этапа').grid(row=0, column=1)



program = PsyProgram()
program.start()
